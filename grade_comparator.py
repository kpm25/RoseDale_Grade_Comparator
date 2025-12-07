import pandas as pd
import re
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np
import os


# --- GLOBAL HELPER FUNCTIONS ---

def get_file_path(file_name):
    """Ensures the file name has the .xlsx extension if it's missing."""
    if not file_name.lower().endswith('.xlsx'):
        return file_name + '.xlsx'
    return file_name


def extract_date_from_sheet_name(sheet_name):
    """Extracts the grade snapshot date (MM-DD-YYYY) from the sheet name."""
    match = re.search(r'(\d{1,2}-\d{1,2}-\d{4})', sheet_name)
    if match:
        return pd.to_datetime(match.group(1), format='%m-%d-%Y', errors='coerce')
    return None


def extract_date_from_file_path(file_path):
    """Extracts the grade snapshot date (D/DDMonYYYY) from the file name."""
    # Updated regex to match 1 or 2 digits for the day (\d{1,2})
    match = re.search(r'(\d{1,2}[A-Za-z]{3}\d{4})', file_path, re.IGNORECASE)
    if match:
        # Removed the deprecated argument infer_datetime_format=True
        return pd.to_datetime(match.group(1), errors='coerce')
    return None


def extract_class_name(file_path):
    """Extracts the course code (e.g., MTH1Wa) from the file name."""
    match = re.search(r'SHEN-([A-Za-z0-9]+)_grades', file_path, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None


def find_completed_assessment_count(df):
    """
    Finds the 'Graded /XX' column, and extracts the first non-null count
    from the data values, which represents the number of completed assignments.
    """
    graded_col_header = None
    for col in df.columns:
        # Step 1: Find the header (e.g., 'Graded /20')
        if re.search(r'Graded\s*/\s*\d+', str(col)):
            graded_col_header = col
            break

    if graded_col_header is None:
        return None

    # Step 2: Get the first non-NaN value in that column (the actual assessment count)
    # This works because the count is uniform across all students in a snapshot.
    first_valid_count = df[graded_col_header].dropna().iloc[0]

    return int(first_valid_count)


def validate_comparison_order(df_old, df_new):
    """
    Validates that the number of completed assessments in the older file
    is less than or equal to the number in the newer file.
    """
    count_old = find_completed_assessment_count(df_old)
    count_new = find_completed_assessment_count(df_new)

    if count_old is None or count_new is None:
        print("\n❌ COMPARISON ERROR: Could not find the 'Graded /XX' column header in one or both files.")
        return False

    if count_old > count_new:
        print("\n❌ COMPARISON ERROR: Invalid timeline.")
        print(f"The older snapshot has {count_old} graded activities, but the newer snapshot has only {count_new}.")
        print("Please ensure the older file is truly chronologically before the newer file.")
        return False

    return True


# ----------------------------------------------------------------------------------

def compare_student_grades(file_path_1, file_path_2):
    """
    Compares student grades between two Excel files, determines the most improved,
    and generates a new, dynamically named, formatted Excel report.
    Returns True on success, False on failure.
    """

    def load_and_clean_data(file_path):
        """Loads data, extracts date, cleans 'Course grade', and ensures all are scaled as percentages."""
        excel_file = pd.ExcelFile(file_path)
        sheet_name = excel_file.sheet_names[0]
        df = pd.read_excel(file_path, sheet_name=sheet_name)

        # 1. Try date extraction
        date = extract_date_from_sheet_name(sheet_name)
        if pd.isna(date):
            date = extract_date_from_file_path(file_path)

        if pd.isna(date):
            raise ValueError(f"Could not extract a valid date from the file: {file_path}")

        # Clean 'Course grade' column: remove '%'
        df['Course grade'] = df['Course grade'].astype(str).str.replace('%', '', regex=False)
        grade_series = pd.to_numeric(df['Course grade'], errors='coerce')

        # *** ULTIMATE SCALING FIX: Scale individual values that look like decimals ***
        is_decimal = (grade_series > 0.0) & (grade_series < 1.5)

        # Apply multiplication only to those specific values
        grade_series.loc[is_decimal] = grade_series.loc[is_decimal] * 100

        df['Course grade'] = grade_series

        return df, date, sheet_name

    # --- 1. Load Data and Extract Dates ---
    try:
        df1, date1, sheet_name1 = load_and_clean_data(file_path_1)
        df2, date2, sheet_name2 = load_and_clean_data(file_path_2)

    except Exception as e:
        print(f"❌ An error occurred during file loading: {e}")
        return False

        # --- 2. Determine Earlier/Later File and Rename Columns ---
    if date1 < date2:
        df_old, df_new = df1, df2
        date_new = date2
    elif date2 < date1:
        df_old, df_new = df2, df1
        date_new = date1
    else:
        print("⚠️ Warning: Both files have the same snapshot date. Comparison aborted.")
        return False

        # *** VALIDATION CHECK ***
    if not validate_comparison_order(df_old, df_new):
        return False

        # Dynamic Output File Naming
    class_name = df_new.iloc[0][df_new.columns[1]]
    class_name = class_name.replace('/', '-')
    output_file = f"{class_name}_Grade_Comparison_Report_{date_new.strftime('%d%b%Y')}.xlsx"
    print(f"🔍 Older snapshot date: {date1.strftime('%m-%d-%Y')}. Newer snapshot date: {date2.strftime('%m-%d-%Y')}.")

    # Logic to identify the dynamic 'Graded/...' column headers
    def find_graded_col(df):
        for col in df.columns:
            if 'Graded' in str(col) and '/' in str(col):
                return col
        return None

    old_graded_col = find_graded_col(df_old)
    new_graded_col = find_graded_col(df_new)

    if not old_graded_col or not new_graded_col:
        print(f"❌ Error: Could not find the 'Graded/...' column in one or both files.")
        return False

        # Select and rename columns for clarity before merging
    df_old = df_old.rename(columns={'Course grade': 'Previous Course Grade (%)'})
    df_new = df_new.rename(columns={'Course grade': 'Current Course Grade (%)',
                                    new_graded_col: 'Current Graded Assessments'})

    # --- 3. Merge Data and Calculate Grade Change ---
    merge_key = df_old.columns[2]
    cols_to_keep_old = [df_old.columns[0], df_old.columns[1], df_old.columns[2], 'Previous Course Grade (%)']
    df_old_data = df_old[cols_to_keep_old]
    cols_to_keep_new = [df_new.columns[2], 'Current Course Grade (%)', 'Current Graded Assessments']
    df_new_data = df_new[cols_to_keep_new]

    merged_df = pd.merge(df_old_data, df_new_data, on=merge_key, how='inner')
    merged_df['Grade Change (%)'] = (
            merged_df['Current Course Grade (%)'] - merged_df['Previous Course Grade (%)']
    )

    # --- 4. Final Data Structure and Most Improved Student(s) (INCLUDING DECLINE) ---
    final_columns = [
        df_old.columns[0], df_old.columns[1], df_old.columns[2],
        'Current Graded Assessments', 'Previous Course Grade (%)',
        'Current Course Grade (%)', 'Grade Change (%)'
    ]
    final_df = merged_df[final_columns].copy()

    # Find MOST IMPROVED (MAX CHANGE)
    max_change = final_df['Grade Change (%)'].max()
    TINY_CHANGE_LIMIT = 0.01

    if max_change > TINY_CHANGE_LIMIT:
        most_improved_students = final_df[final_df['Grade Change (%)'] == max_change][df_old.columns[2]].tolist()
    else:
        not_significantly_worse = final_df[final_df['Grade Change (%)'] >= -TINY_CHANGE_LIMIT]
        if not not_significantly_worse.empty:
            max_current_grade = not_significantly_worse['Current Course Grade (%)'].max()
            most_improved_students = not_significantly_worse[
                not_significantly_worse['Current Course Grade (%)'] == max_current_grade
                ][df_old.columns[2]].tolist()
        else:
            most_improved_students = final_df[final_df['Grade Change (%)'] == max_change][df_old.columns[2]].tolist()

    final_df['Most Improved Student(s)'] = final_df[final_df.columns[2]].apply(
        lambda x: '🥇 MOST IMPROVED' if x in most_improved_students else ''
    )

    # *** ADDED: Find MOST DECLINED (MIN CHANGE) ***
    min_change = final_df['Grade Change (%)'].min()

    most_declined_students = final_df[final_df['Grade Change (%)'] == min_change][df_old.columns[2]].tolist()

    # Column I: Biggest Decline (Moved from H to I to accommodate the new column J)
    final_df['Biggest Decline'] = final_df[final_df.columns[2]].apply(
        lambda x: '🔻 BIGGEST DROP' if x in most_declined_students else ''
    )

    # --- 5. Output to Formatted Excel ---

    try:
        # Save the unformatted data to a temporary Excel writer
        writer = pd.ExcelWriter(output_file, engine='openpyxl')
        final_df.to_excel(writer, index=False, sheet_name='Grade Report')
        writer.close()

        # Reload the workbook to apply conditional formatting and styling
        wb = load_workbook(output_file)
        ws = wb['Grade Report']

        # --- Column Width Adjustment ---
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        # --- End Column Width Adjustment ---

        # --- Code for Column Centering (D to I) ---
        center_alignment = Alignment(horizontal='center', vertical='center')
        # Center Columns D (index 4) through J (index 10)
        for col_idx in range(4, 11):  # <-- Range adjusted to cover new Column I (index 9) and J (index 10)
            for row_idx in range(1, ws.max_row + 1):
                ws.cell(row=row_idx, column=col_idx).alignment = center_alignment
        # --- End Code for Column Centering ---

        # Define fill colors
        fill_green = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        fill_red = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
        fill_cyan = PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid")
        fill_yellow = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
        # *** ADDED NEW COLOR ***
        fill_purple = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")

        for row_idx in range(2, ws.max_row + 1):
            grade_change_cell = ws.cell(row=row_idx, column=7)  # Column G: Grade Change
            winner_status_cell = ws.cell(row=row_idx, column=8)  # Column H: Most Improved
            decline_status_cell = ws.cell(row=row_idx, column=9)  # Column I: Biggest Decline (New)
            grade_change_value = grade_change_cell.value

            if pd.isna(grade_change_value): continue

            # --- 1. WINNER IS ALWAYS GREEN (Highest Priority) ---
            if winner_status_cell.value == '🥇 MOST IMPROVED':
                grade_change_cell.fill = fill_green

                # Apply '0.00' format IF the change is zero
                if abs(grade_change_value) <= TINY_CHANGE_LIMIT:
                    grade_change_cell.value = 0.0
                    grade_change_cell.number_format = '0.00'

            # *** ADDED: BIGGEST DROP IS PURPLE ***
            elif decline_status_cell.value == '🔻 BIGGEST DROP':
                grade_change_cell.fill = fill_purple

            # --- 2. OTHER NO-CHANGE STUDENTS ARE YELLOW ---
            elif abs(grade_change_value) <= TINY_CHANGE_LIMIT:
                grade_change_cell.fill = fill_yellow
                grade_change_cell.value = 0.0
                grade_change_cell.number_format = '0.00'

            # --- 3. NEGATIVE CHANGE IS RED ---
            elif grade_change_value < 0:
                grade_change_cell.fill = fill_red

            # --- 4. POSITIVE (NON-WINNING) CHANGE IS CYAN ---
            elif grade_change_value > 0:
                grade_change_cell.fill = fill_cyan

        wb.save(output_file)
        print(f"\n✅ Success! Report generated and saved as '{output_file}'")
        return True  # SUCCESS EXIT

    except PermissionError:
        print("\n❌ ERROR: Permission denied.")
        print(
            f"Please ensure the output file '{output_file}' is CLOSED and not open in Excel or any other program, then run the script again.")
        return False  # FAILURE EXIT 5
    except Exception as e:
        print(f"\n❌ An unexpected error occurred during saving: {e}")
        return False  # FAILURE EXIT 6
    finally:
        # UNIVERSAL FIX: Pause the console window before closing
        print("\n\n---------------------------------------------")
        input("✅ Processing complete. Press Enter to exit and view the output file.")


# --- EXECUTION ---

# Define the file names used for the default option
DEFAULT_FILE_1 = 'SHEN-MTH1Wa_grades_28Nov2025.xlsx'
DEFAULT_FILE_2 = 'SHEN-MTH1Wa_grades_30Nov2025.xlsx'


def get_file_paths():
    """Presents an interactive menu to get file paths from the user and validates class names."""
    print("\n=============================================")
    print("           GRADE COMPARISON TOOL")
    print("=============================================")
    print("ℹ️ NOTE: Ensure the Excel files are in the same folder as this application.")
    print("---------------------------------------------")
    print(f"1. Use Default Files: ({DEFAULT_FILE_1} & {DEFAULT_FILE_2})")
    print("2. Enter File Names Manually (no need to type .xlsx)")
    print("---------------------------------------------")

    max_retries = 3
    retries = 0

    while retries < max_retries:
        choice = input("Enter choice (1 or 2): ").strip()

        if choice == '1':
            file1, file2 = DEFAULT_FILE_1, DEFAULT_FILE_2
            if os.path.exists(file1) and os.path.exists(file2):

                # Validation
                class1 = extract_class_name(file1)
                class2 = extract_class_name(file2)

                if class1 and class2 and class1 == class2:
                    print(f"✅ Using default files.")
                    return file1, file2
                else:
                    print("\n❌ ERROR: Default files class mismatch or not properly named. RETRYING...")
                    retries += 1
                    continue

            else:
                print("\n❌ ERROR: Default files not found in the current directory. RETRYING...")
                retries += 1
                continue

        elif choice == '2':
            # --- CHANGE 1: Streamline Prompts ---
            raw_input_1 = input("Enter OLDER snapshot file name: ").strip()
            raw_input_2 = input("Enter NEWER snapshot file name: ").strip()

            file1 = get_file_path(raw_input_1)
            file2 = get_file_path(raw_input_2)

            if os.path.exists(file1) and os.path.exists(file2):
                # Validation
                class1 = extract_class_name(file1)
                class2 = extract_class_name(file2)

                if not class1 or not class2:
                    print(
                        "\n❌ ERROR: Could not identify class name (e.g., MTH1Wa) in one or both file names. RETRYING...")
                    retries += 1
                    continue

                if class1 == class2:
                    print(f"✅ Class match ({class1}). Files found.")
                    return file1, file2
                else:
                    print("\n❌ ERROR: Class mismatch! Please enter files for the same course. RETRYING...")
                    retries += 1
                    continue
            else:
                print("\n❌ ERROR: One or both files were not found. RETRYING...")
                retries += 1
                continue

        else:
            print("Invalid choice. Please enter 1 or 2. RETRYING...")
            retries += 1
            continue

    print(f"\n🚫 Maximum {max_retries} retries reached. Terminating program.")
    return None, None


# --- MAIN PROGRAM LOOP ---

def main():
    while True:
        # 1. Get file paths and run comparison
        file_1_path, file_2_path = get_file_paths()

        if file_1_path is not None:
            # Only run comparison if valid files were returned
            compare_student_grades(file_1_path, file_2_path)

        # The compare_student_grades function contains the final input() pause,
        # so this loop restarts after the user hits Enter inside that function.

        # 2. Add "Run Again" Option (CHANGE 2)
        print("\n\n*********************************************")
        choice = input("Do you want to run another file comparison? (Y/N): ").strip().upper()
        print("*********************************************")

        if choice != 'Y':
            # This is the graceful exit point for the entire application
            print("\nTool terminated successfully. Goodbye!")
            break


if __name__ == '__main__':
    main()